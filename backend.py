import json
import time
from datetime import datetime, timedelta
import pandas as pd
import csv
from openpyxl import load_workbook

class DbAct:
    def __init__(self, db, config, path_xlsx):
        super(DbAct, self).__init__()
        self.__db = db
        self.__config = config
        self.__fields_user = ['Имя', 'Фамилия', 'Никнейм', 'Номер телефона']
        self.__dump_path_xlsx = path_xlsx


    def add_user(self, user_id, full_name, nick_name, is_foreman):
            if not self.user_is_existed(user_id):
                if user_id in self.__config.get_config()['admins']:
                    is_admin = True
                else:
                    is_admin = False
                self.__db.db_write(
                    'INSERT INTO users (user_id, full_name, nick_name, system_data, is_admin, is_foreman) '
                    'VALUES (?, ?, ?, ?, ?, ?)',
                    (user_id, full_name, nick_name, json.dumps({"index": None, 
                                                                "attach_object_name": None,
                                                                "full_name": None,
                                                                "object_id": None,
                                                                "admin_object_id": None,
                                                                "category_id": None,
                                                                "subcategory_id": None,
                                                                "work_type_id": None,
                                                                "work_type_name": None,
                                                                "work_type_unit": None,
                                                                "work_type_volume": None,
                                                                "material_id": None,
                                                                "material_date": None,
                                                                "material_name": None,
                                                                "material_norm": None,
                                                                "material_unit": None,
                                                                "material_counterparty": None,
                                                                "material_registration_number": None,
                                                                "material_volume": None,
                                                                "technique_name": None,
                                                                "technique_contragent": None,
                                                                "technique_number": None,
                                                                "technique_unit": None,
                                                                "technique_volume": None,
                                                                "coming_date": None,
                                                                "coming_name": None,
                                                                "coming_unit": None,
                                                                "coming_volume": None,
                                                                "coming_supplier": None}), is_admin, is_foreman))
                
    def set_user_id_in_topic(self, user_id, topic_id):
        if not self.user_is_existed(user_id):
            return None
        self.__db.db_write('UPDATE users SET topic_id = ? WHERE user_id = ?', (topic_id, user_id))
                
    def get_user_id_from_topic(self, topic_id):
        data = self.__db.db_read("SELECT user_id FROM users WHERE topic_id = ?", (topic_id, ))
        if len(data) > 0:
            return data[0][0]
        
    def set_user_is_foreman(self, user_id):
        if not self.user_is_existed(user_id):
            return None
        self.__db.db_write('UPDATE users SET is_foreman = 1 WHERE user_id = ?', (user_id, ))

    def set_user_full_name(self, user_id, full_name):
        if not self.user_is_existed(user_id):
            return None
        self.__db.db_write('UPDATE users SET full_name = ? WHERE user_id = ?', (full_name, user_id,))


    def user_is_existed(self, user_id: int):
            data = self.__db.db_read('SELECT count(*) FROM users WHERE user_id = ?', (user_id,))
            if len(data) > 0:
                if data[0][0] > 0:
                    status = True
                else:
                    status = False
                return status
            
    def user_is_admin(self, user_id: int):
        data = self.__db.db_read('SELECT is_admin FROM users WHERE user_id = ?', (user_id,))
        if len(data) > 0:
            if data[0][0] == 1:
                status = True
            else:
                status = False
            return status
        
    def user_is_foreman(self, user_id: int):
        data = self.__db.db_read('SELECT is_foreman FROM users WHERE user_id = ?', (user_id,))
        if len(data) > 0:
            if data[0][0] == 1:
                status = True
            else:
                status = False
            return status
        
    def set_user_system_key(self, user_id: int, key: str, value: any) -> None:
        system_data = self.get_user_system_data(user_id)
        if system_data is None:
            return None
        system_data = json.loads(system_data)
        system_data[key] = value
        self.__db.db_write('UPDATE users SET system_data = ? WHERE user_id = ?', (json.dumps(system_data), user_id))

    def get_user_system_key(self, user_id: int, key: str):
        system_data = self.get_user_system_data(user_id)
        if system_data is None:
            return None
        system_data = json.loads(system_data)
        if key not in system_data.keys():
            return None
        return system_data[key]

    def get_user_system_data(self, user_id: int):
        if not self.user_is_existed(user_id):
            return None
        return self.__db.db_read('SELECT system_data FROM users WHERE user_id = ?', (user_id,))[0][0]
    
    def write_new_object(self, user_id, object_name):
        if not self.user_is_existed(user_id):
            return None
        self.__db.db_write('INSERT INTO construction_objects (object_name) VALUES (?)', (object_name,))

    def get_all_objects(self, user_id):
        if not self.user_is_existed(user_id):
            return None
        return self.__db.db_read('SELECT user_id, object_name FROM construction_objects', ())
    
    def get_admin_objects(self, user_id):
        if not self.user_is_existed(user_id):
            return None
        return self.__db.db_read('SELECT row_id, object_name FROM construction_objects', ())

    def get_foreman_objects(self, user_id):
        if not self.user_is_existed(user_id):
            return None
        return self.__db.db_read('SELECT row_id, object_name FROM construction_objects WHERE user_id = ?', (user_id,))

    def get_name_from_user_id(self, user_id):
        if not self.user_is_existed(user_id):
            return None
        return self.__db.db_read('SELECT full_name FROM users WHERE user_id = ?', (user_id,))
    
    def get_foremans(self, user_id):
        if not self.user_is_existed(user_id):
            return None
        return self.__db.db_read('SELECT user_id, full_name FROM users WHERE is_foreman = True', ())
    
    def attach_foreman_to_object(self, user_id, object_name):
        if not self.user_is_existed(user_id):
            return None
        self.__db.db_write('UPDATE construction_objects SET user_id = ? WHERE object_name = ?', (user_id, object_name,))
    
    def unpin_foreman_from_object(self, user_id, object_name):
        if not self.user_is_existed(user_id):
            return None
        self.__db.db_write('UPDATE construction_objects SET user_id = ? WHERE object_name = ?', (None, object_name,))
    
    def delete_object(self, user_id, object_name):
        if not self.user_is_existed(user_id):
            return None
        self.__db.db_write('DELETE FROM construction_objects WHERE object_name = ?', (object_name, ))
    

    def add_work_categories(self, user_id, object_id, name):
        if not self.user_is_existed(user_id):
            return None
        self.__db.db_write('INSERT INTO work_categories (object_id, name) VALUES (?, ?)', (object_id, name,))

    def get_work_categories(self, user_id, object_id):
        if not self.user_is_existed(user_id):
            return None
        return self.__db.db_read('SELECT row_id, name FROM work_categories WHERE object_id = ?', (object_id,))

    def delete_work_categories(self, user_id, category_id):
        if not self.user_is_existed(user_id):
            return None
        self.__db.db_write('DELETE FROM work_categories WHERE row_id = ?', (category_id,))

    def add_work_subcategories(self, user_id, category_id, name):
        if not self.user_is_existed(user_id):
            return None
        self.__db.db_write('INSERT INTO work_subcategories (category_id, name) VALUES (?, ?)', (category_id, name,))
    
    def get_work_subcategories(self, user_id, category_id):
        if not self.user_is_existed(user_id):
            return None
        return self.__db.db_read('SELECT row_id, name FROM work_subcategories WHERE category_id = ?', (category_id,))

    def delete_work_subcategories(self, user_id, subcategory_id):
        if not self.user_is_existed(user_id):
            return None
        self.__db.db_write('DELETE FROM work_subcategories WHERE row_id = ?', (subcategory_id,))

    def add_work_type(self, user_id, subcategory_id, name, unit):
        if not self.user_is_existed(user_id):
            return None
        self.__db.db_write('INSERT INTO work_types (subcategory_id, name, unit) VALUES (?, ?, ?)', (subcategory_id, name, unit,))
    
    def get_work_type(self, user_id, subcategory_id):
        if not self.user_is_existed(user_id):
            return None
        return self.__db.db_read('SELECT row_id, name FROM work_types WHERE subcategory_id = ?', (subcategory_id,))
    
    def delete_work_type(self, user_id, work_type_id):
        if not self.user_is_existed(user_id):
            return None
        self.__db.db_write('DELETE FROM work_types WHERE row_id = ?', (work_type_id,))

    def add_work_material(self, user_id, work_type_id, date, name, norm, unit, counterparty, registration_number, volume, cost):
        if not self.user_is_existed(user_id):
            return None
        self.__db.db_write('INSERT INTO work_materials (work_type_id, date, name, norm, unit, counterparty, state_registration_number_vehicle, volume, cost) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)', (work_type_id, date, name, norm, unit, counterparty, registration_number, volume, cost,))

    def get_work_material(self, user_id, work_type_id):
        if not self.user_is_existed(user_id):
            return None
        return self.__db.db_read('SELECT row_id, name, norm, unit, counterparty, state_registration_number_vehicle, volume, cost FROM work_materials WHERE work_type_id = ?', (work_type_id,))
    
    def delete_work_material(self, user_id, row_id):
        if not self.user_is_existed(user_id):
            return None
        self.__db.db_write('DELETE FROM work_materials WHERE row_id = ?', (row_id,))

    def add_technique(self, user_id, object_id, work_type_id, technique_name, technique_contagent, technique_number, technique_unit, technique_volume, technique_cost):
        if not self.user_is_existed(user_id):
            return None
        self.__db.db_write('INSERT INTO list_technique (object_id, work_type_id, name, counterparty, state_registration_number_vehicle, unit, volume, cost) VALUES (?, ?, ?, ?, ?, ?, ?, ?)', (object_id, work_type_id, technique_name, technique_contagent, technique_number, technique_unit, technique_volume, technique_cost,))
    
    def delete_technique(self, user_id, row_id):
        if not self.user_is_existed(user_id):
            return None
        self.__db.db_write('DELETE FROM list_technique WHERE row_id = ?', (row_id))

    def get_list_technique(self, user_id, object_id):
        if not self.user_is_existed(user_id):
            return None
        return self.__db.db_read('SELECT row_id, name FROM list_technique WHERE object_id = ?', (object_id,))

    def add_list_coming(self, user_id, object_id, date, name, unit, volume, supplier, cost):
        if not self.user_is_existed(user_id):
            return None
        self.__db.db_write('INSERT INTO list_coming (object_id, date, name, unit, volume, supplier, cost) VALUES (?, ?, ?, ?, ?, ?, ?)', (object_id, date, name, unit, volume, supplier, cost))

    def get_from_list_coming(self, user_id, object_id):
        if not self.user_is_existed(user_id):
            return None
        return self.__db.db_read('SELECT row_id, name FROM list_coming WHERE object_id = ?', (object_id,))

    def delete_from_list_coming(self, user_id, row_id):
        if not self.user_is_existed(user_id):
            return None
        self.__db.db_write('DELETE FROM list_coming WHERE row_id = ?', (row_id,))

    def add_mateials_norm(self, user_id, norm, row_id):
        if not self.user_is_existed(user_id):
            return None
        self.__db.db_write('UPDATE work_materials SET norm = ? WHERE row_id = ?', (norm, row_id,))

    def add_materials_unit(self, user_id, unit, row_id):
        if not self.user_is_existed(user_id):
            return None
        self.__db.db_write('UPDATE work_materials SET unit = ? WHERE row_id = ?', (unit, row_id,))
    
    def add_materials_number(self, user_id, number, row_id):
        if not self.user_is_existed(user_id):
            return None
        self.__db.db_write('UPDATE work_materials SET state_registration_number_vehicle = ? WHERE row_id = ?', (number, row_id,))

    def select_material_smr(self, user_id, work_type_id):
        if not self.user_is_existed(user_id):
            return None
        return self.__db.db_read('SELECT name = ? FROM work_materials WHERE work_type_id = ?', ("СМР", work_type_id,))
    
    def edit_material_smr(self, user_id, unit, volume, cost, work_type_id):
        if not self.user_is_existed(user_id):
            return None
        self.__db.db_write('UPDATE work_materials SET unit = ?, volume = ?, cost = ? WHERE work_type_id = ?', (unit, volume, cost, work_type_id))
    
    def add_material_smr(self, user_id, date, volume, unit, cost, work_type_id):
        if not self.user_is_existed(user_id):
            return None
        self.__db.db_write('INSERT INTO work_materials (date, work_type_id, name, unit, volume, cost) VALUES (?, ?, "СМР", ?, ?, ?)', (date, work_type_id, unit, volume, cost,))
    
    def db_export_object_report(self, object_id):
        """Экспортирует данные объекта строительства в XLSX-файл с правильной структурой сумм"""
        try:
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from collections import defaultdict
            import os
            from datetime import datetime

            # Стили оформления
            styles = {
                'object_name': PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid'),
                'header': PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid'),
                'category': PatternFill(start_color='548235', end_color='548235', fill_type='solid'),
                'subcategory': PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid'),
                'work_type': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
                'material_font': Font(color='0070C0'),
                'technique': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
                'technique_font': Font(color='000000', italic=True),
                'header_font': Font(bold=True, color='FFFFFF'),
                'category_font': Font(bold=True, color='FFFFFF'),
                'subcategory_font': Font(bold=True, color='000000'),
                'border': Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin')),
                'number_format': '0.0000',
                'money_format': '#,##0.00',
                'date_format': 'YYYY-MM-DD'
            }

            def safe_float(value):
                """Преобразует значение в float с обработкой русских форматов"""
                try:
                    if value is None or value == '':
                        return 0.0
                    if isinstance(value, str):
                        value = value.replace(' ', '').replace(',', '.')
                        cleaned = ''.join([c for c in value if c.isdigit() or c == '.'])
                        if not cleaned:
                            return 0.0
                        return round(float(cleaned), 4)
                    return round(float(value), 4)
                except (ValueError, TypeError) as e:
                    print(f"Ошибка преобразования '{value}': {e}")
                    return 0.0

            def safe_str(value):
                """Безопасное преобразование в строку"""
                if value is None:
                    return ''
                return str(value).strip()

            report_filename = "report.xlsx"
            
            object_data = self.__db.db_read(
                'SELECT row_id, object_name FROM construction_objects WHERE row_id = ?', 
                (object_id,))
            if not object_data:
                print(f"Объект с ID {object_id} не найден")
                return False
                    
            obj_id, obj_name = object_data[0]

            # Создаем новую книгу Excel
            from openpyxl import Workbook
            wb = Workbook()
            wb.remove(wb.active)

            # Создаем листы
            ws_prorab = wb.create_sheet("Прораб")
            ws_material = wb.create_sheet("Материал-Работа")
            ws_tech = wb.create_sheet("Техника")
            ws_coming = wb.create_sheet("Приход")

            # ==================== ЛИСТ "ПРОРАБ" ====================
            works_data = []
            all_materials = []
            technique_data_by_work = defaultdict(list)  # Для группировки техники по work_type_id
            
            # Заголовки
            columns = [
                '№ п/п', 'Наименование работ/материала', 'норма', 'ед.изм.',
                'контрагент', 'гос.№ (техники)', 'объем', 'цена', 'сумма'
            ]
            
            works_data.append([f"Наименование объекта: {obj_name}"] + [None]*7 + [0.0])
            works_data.append([None]*9)
            works_data.append(columns)
            
            # Получаем категории работ
            categories = self.__db.db_read(
                'SELECT row_id, name FROM work_categories WHERE object_id = ? ORDER BY row_id',
                (obj_id,))
            
            # Получаем данные о технике и группируем их по work_type_id
            techniques = self.__db.db_read(
                'SELECT row_id, work_type_id, name, counterparty, state_registration_number_vehicle, unit, volume, cost '
                'FROM list_technique WHERE object_id = ?',
                (obj_id,))
            
            for tech in techniques:
                tech_id, work_type_id, name, counterparty, reg_number, unit, volume, cost = tech
                technique_data_by_work[work_type_id].append({
                    'name': safe_str(name),
                    'counterparty': safe_str(counterparty),
                    'reg_number': safe_str(reg_number),
                    'unit': safe_str(unit),
                    'volume': safe_float(volume),
                    'cost': safe_float(cost),
                    'sum': safe_float(volume) * safe_float(cost)
                })

            # Словари для хранения сумм
            category_sums = defaultdict(float)
            subcategory_sums = defaultdict(float)
            work_type_sums = defaultdict(float)
            
            # Обрабатываем категории, подкатегории и работы
            for cat_idx, (cat_id, cat_name) in enumerate(categories, 1):
                works_data.append([f"{cat_idx}", cat_name] + [None]*7 + [0.0])
                current_category_row = len(works_data) - 1
                
                subcategories = self.__db.db_read(
                    'SELECT row_id, name FROM work_subcategories WHERE category_id = ? ORDER BY row_id',
                    (cat_id,))
                
                for sub_idx, (subcat_id, subcat_name) in enumerate(subcategories, 1):
                    works_data.append([f"{cat_idx}.{sub_idx}", subcat_name] + [None]*7 + [0.0])
                    current_subcategory_row = len(works_data) - 1
                    
                    work_types = self.__db.db_read(
                        'SELECT row_id, name, unit FROM work_types WHERE subcategory_id = ? ORDER BY row_id',
                        (subcat_id,))
                    
                    for wt_idx, (wt_id, wt_name, wt_unit) in enumerate(work_types, 1):
                        # Получаем материалы для работы
                        materials = self.__db.db_read(
                            'SELECT date, name, norm, unit, counterparty, state_registration_number_vehicle, volume, cost '
                            'FROM work_materials WHERE work_type_id = ? ORDER BY date',
                            (wt_id,))
                        
                        total_materials_sum = 0.0
                        material_groups = defaultdict(list)
                        smr_volume = 0.0
                        
                        # Обрабатываем материалы
                        for mat in materials:
                            try:
                                volume = safe_float(mat[6])
                                cost = safe_float(mat[7])
                                mat_sum = volume * cost
                                total_materials_sum += mat_sum
                                
                                mat_name = safe_str(mat[1])
                                material_groups[mat_name.lower()].append({
                                    'name': mat_name,
                                    'norm': safe_str(mat[2]),
                                    'unit': safe_str(mat[3]),
                                    'counterparty': safe_str(mat[4]),
                                    'reg_number': safe_str(mat[5]),
                                    'volume': volume,
                                    'cost': cost,
                                    'sum': mat_sum
                                })
                                
                                if mat_name.lower() == 'смр':
                                    smr_volume = volume
                                
                                all_materials.append({
                                    'date': mat[0],
                                    'name': mat_name,
                                    'norm': safe_str(mat[2]),
                                    'unit': safe_str(mat[3]),
                                    'counterparty': safe_str(mat[4]),
                                    'state_reg_number': safe_str(mat[5]),
                                    'volume': volume,
                                    'cost': cost,
                                    'sum': mat_sum,
                                    'work_id': f"{cat_idx}.{sub_idx}.{wt_idx}"
                                })
                            except Exception:
                                continue
                        
                        # Добавляем технику к общей сумме
                        total_technique_sum = 0.0
                        if wt_id in technique_data_by_work:
                            for tech in technique_data_by_work[wt_id]:
                                total_technique_sum += tech['sum']
                        
                        total_work_sum = total_materials_sum + total_technique_sum
                        
                        # Сохраняем суммы для агрегации
                        work_type_sums[wt_id] = total_work_sum
                        subcategory_sums[subcat_id] += total_work_sum
                        category_sums[cat_id] += total_work_sum
                        
                        # Добавляем вид работы с общей суммой
                        works_data.append([
                            f"{cat_idx}.{sub_idx}.{wt_idx}",
                            wt_name,
                            None,
                            wt_unit,
                            None, None,
                            smr_volume,
                            total_work_sum / smr_volume if smr_volume != 0 else 0,
                            total_work_sum
                        ])
                        current_work_type_row = len(works_data) - 1
                        
                        # Добавляем материалы (с суммированием)
                        for mat_group in material_groups.values():
                            if not mat_group:
                                continue
                                
                            combined = {
                                'name': mat_group[0]['name'],
                                'norms': set(),
                                'units': set(),
                                'counterparties': set(),
                                'reg_numbers': set(),
                                'volumes': [],
                                'costs': [],
                                'sum': 0.0
                            }
                            
                            for mat in mat_group:
                                if mat['norm']: combined['norms'].add(mat['norm'])
                                if mat['unit']: combined['units'].add(mat['unit'])
                                if mat['counterparty']: combined['counterparties'].add(mat['counterparty'])
                                if mat['reg_number']: combined['reg_numbers'].add(mat['reg_number'])
                                combined['volumes'].append(mat['volume'])
                                combined['costs'].append(mat['cost'])
                                combined['sum'] += mat['sum']
                            
                            works_data.append([
                                None,
                                combined['name'],
                                ', '.join(filter(None, combined['norms'])) if combined['norms'] else None,
                                ', '.join(filter(None, combined['units'])) if combined['units'] else None,
                                ', '.join(filter(None, combined['counterparties'])) if combined['counterparties'] else None,
                                ', '.join(filter(None, combined['reg_numbers'])) if combined['reg_numbers'] else None,
                                sum(combined['volumes']),
                                sum(combined['costs'])/len(combined['costs']) if combined['costs'] else 0,
                                combined['sum']
                            ])
                        
                        # Добавляем технику (с суммированием)
                        if wt_id in technique_data_by_work:
                            for tech in technique_data_by_work[wt_id]:
                                works_data.append([
                                    None,
                                    tech['name'],
                                    None,
                                    tech['unit'],
                                    tech['counterparty'],
                                    tech['reg_number'],
                                    tech['volume'],
                                    tech['cost'],
                                    tech['sum']
                                ])
                        
                        # Обновляем сумму вида работы
                        works_data[current_work_type_row][8] = total_work_sum
                        works_data[current_work_type_row][7] = total_work_sum / smr_volume if smr_volume != 0 else 0
                    
                    # Обновляем сумму подкатегории
                    works_data[current_subcategory_row][8] = subcategory_sums[subcat_id]
                
                # Обновляем сумму категории
                works_data[current_category_row][8] = category_sums[cat_id]
            
            # Обновляем общую сумму объекта
            works_data[0][8] = sum(category_sums.values())

            # Добавляем технику без привязки к работе (если есть)
            techniques_without_work = self.__db.db_read(
                'SELECT name, counterparty, state_registration_number_vehicle, unit, volume, cost '
                'FROM list_technique WHERE object_id = ? AND work_type_id IS NULL',
                (obj_id,))
            
            tech_groups = defaultdict(list)
            for tech in techniques_without_work:
                tech_sum = safe_float(tech[4]) * safe_float(tech[5])
                tech_groups[tech[0].lower()].append({
                    'name': tech[0],
                    'counterparty': tech[1],
                    'reg_number': tech[2],
                    'unit': tech[3],
                    'volume': safe_float(tech[4]),
                    'cost': safe_float(tech[5]),
                    'sum': tech_sum
                })
            
            for tech_group in tech_groups.values():
                combined = {
                    'name': tech_group[0]['name'],
                    'counterparties': set(),
                    'reg_numbers': set(),
                    'units': set(),
                    'volumes': [],
                    'costs': [],
                    'sum': 0.0
                }
                
                for tech in tech_group:
                    if tech['counterparty']: combined['counterparties'].add(tech['counterparty'])
                    if tech['reg_number']: combined['reg_numbers'].add(tech['reg_number'])
                    if tech['unit']: combined['units'].add(tech['unit'])
                    combined['volumes'].append(tech['volume'])
                    combined['costs'].append(tech['cost'])
                    combined['sum'] += tech['sum']
                
                works_data.append([
                    None,
                    combined['name'],
                    None,
                    ', '.join(filter(None, combined['units'])) if combined['units'] else None,
                    ', '.join(filter(None, combined['counterparties'])) if combined['counterparties'] else None,
                    ', '.join(filter(None, combined['reg_numbers'])) if combined['reg_numbers'] else None,
                    sum(combined['volumes']),
                    sum(combined['costs'])/len(combined['costs']) if combined['costs'] else 0,
                    combined['sum']
                ])
                # Добавляем сумму техники без привязки к общей сумме объекта
                works_data[0][8] += combined['sum']

            # Записываем данные в лист "Прораб"
            for row in works_data:
                ws_prorab.append(row)
            
            # Применяем стили
            for row in ws_prorab.iter_rows():
                for cell in row:
                    cell.border = styles['border']
                    
                    if cell.row == 1:
                        cell.fill = styles['object_name']
                        cell.font = Font(bold=True)
                    elif cell.row == 3:
                        cell.fill = styles['header']
                        cell.font = styles['header_font']
                    elif cell.row > 3 and isinstance(works_data[cell.row-1][0], str) and '.' not in str(works_data[cell.row-1][0]):
                        cell.fill = styles['category']
                        cell.font = styles['category_font']
                    elif cell.row > 3 and isinstance(works_data[cell.row-1][0], str) and str(works_data[cell.row-1][0]).count('.') == 1:
                        cell.fill = styles['subcategory']
                        cell.font = styles['subcategory_font']
                    elif cell.row > 3 and isinstance(works_data[cell.row-1][0], str) and str(works_data[cell.row-1][0]).count('.') == 2:
                        cell.fill = styles['work_type']
                        cell.font = Font()
                    elif cell.row > 3 and cell.column == 2 and works_data[cell.row-1][0] is None:
                        cell.font = styles['material_font']
                    
                    if isinstance(cell.value, (int, float)):
                        if cell.column in [7, 8]:
                            cell.number_format = styles['number_format']
                        elif cell.column == 9:
                            cell.number_format = styles['money_format']
                    
                    cell.alignment = Alignment(horizontal='left', vertical='center')

            # Настройка ширины столбцов
            for col, width in {'A':8, 'B':40, 'C':10, 'D':8, 'E':15, 'F':15, 'G':12, 'H':12, 'I':12}.items():
                ws_prorab.column_dimensions[col].width = width

            # ==================== ЛИСТ "МАТЕРИАЛ-РАБОТА" ====================
            materials_data = []
            materials_columns = [
                'Дата', 'Наименование материала', 'Норма', 'ед.изм.',
                'Контрагент', 'Гос.№ техники', 'Объем', 'Цена', 'Работа'
            ]
            
            materials_data.append([f"Наименование объекта: {obj_name}"] + [None]*8)
            materials_data.append([None]*9)
            materials_data.append(materials_columns)

            all_materials_sorted = sorted(all_materials, key=lambda x: (x['date'], x['work_id']))
            
            for mat in all_materials_sorted:
                materials_data.append([
                    mat['date'],
                    mat['name'],
                    mat['norm'],
                    mat['unit'],
                    mat['counterparty'],
                    mat['state_reg_number'],
                    mat['volume'],
                    mat['cost'],
                    mat['work_id']
                ])

            for row in materials_data:
                ws_material.append(row)
            
            for row in ws_material.iter_rows():
                for cell in row:
                    cell.border = styles['border']
                    
                    if cell.row == 1:
                        cell.fill = styles['object_name']
                        cell.font = Font(bold=True)
                    elif cell.row == 3:
                        cell.fill = styles['header']
                        cell.font = styles['header_font']
                    elif cell.column == 1 and cell.row > 3:
                        cell.number_format = styles['date_format']
                    
                    if isinstance(cell.value, (int, float)):
                        if cell.column in [7, 8]:
                            cell.number_format = styles['number_format']
                    
                    cell.alignment = Alignment(horizontal='left', vertical='center')

            for col, width in {'A':12, 'B':40, 'C':10, 'D':8, 'E':15, 'F':15, 'G':12, 'H':12, 'I':15}.items():
                ws_material.column_dimensions[col].width = width

            # ==================== ЛИСТ "ТЕХНИКА" ====================
            tech_data = []
            tech_columns = [
                'Наименование техники', 'Контрагент', 'Гос.№ техники', 
                'Ед.изм.', 'Объем (часы)', 'Цена за час', 'Сумма', 'Вид работы'
            ]
            
            # Получаем все данные о технике с названиями видов работ
            techniques = self.__db.db_read('''
                SELECT t.name, t.counterparty, t.state_registration_number_vehicle, 
                    t.unit, t.volume, t.cost, wt.name
                FROM list_technique t
                LEFT JOIN work_types wt ON t.work_type_id = wt.row_id
                WHERE t.object_id = ?
            ''', (obj_id,))
            
            for tech in techniques:
                tech_sum = safe_float(tech[4]) * safe_float(tech[5])
                tech_data.append([
                    tech[0],
                    tech[1],
                    tech[2],
                    tech[3],
                    safe_float(tech[4]),
                    safe_float(tech[5]),
                    tech_sum,
                    tech[6] if tech[6] else 'Не указан'
                ])

            ws_tech.append(tech_columns)
            for row in tech_data:
                ws_tech.append(row)

            for row in ws_tech.iter_rows():
                for cell in row:
                    cell.border = styles['border']
                    
                    if cell.row == 1:
                        cell.fill = styles['header']
                        cell.font = styles['header_font']
                    
                    if isinstance(cell.value, (int, float)):
                        if cell.column in [5, 6]:
                            cell.number_format = styles['number_format']
                        elif cell.column == 7:
                            cell.number_format = styles['money_format']
                    
                    cell.alignment = Alignment(horizontal='left', vertical='center')

            for col, width in {'A':25, 'B':20, 'C':15, 'D':10, 'E':12, 'F':12, 'G':12, 'H':25}.items():
                ws_tech.column_dimensions[col].width = width

            # ==================== ЛИСТ "ПРИХОД" ====================
            coming_data = []
            coming_columns = ['№', 'Дата', 'Наименование', 'Ед.изм.', 'Объем', 'Поставщик', 'Цена', 'ИТОГО', 'ПРИМЕЧАНИЕ']
            
            comings = self.__db.db_read(
                'SELECT date, name, unit, volume, supplier, cost '
                'FROM list_coming WHERE object_id = ? ORDER BY date',
                (obj_id,))
            
            for idx, (date, name, unit, volume, supplier, cost) in enumerate(comings, 1):
                coming_sum = safe_float(volume) * safe_float(cost)
                coming_data.append([
                    idx,
                    datetime.strptime(date, '%Y-%m-%d').date() if isinstance(date, str) else date,
                    name,
                    unit,
                    safe_float(volume),
                    supplier,
                    safe_float(cost),
                    coming_sum,
                    'приход'
                ])

            ws_coming.append(coming_columns)
            for row in coming_data:
                ws_coming.append(row)

            for row in ws_coming.iter_rows():
                for cell in row:
                    cell.border = styles['border']
                    
                    if cell.row == 1:
                        cell.fill = styles['header']
                        cell.font = styles['header_font']
                    elif cell.column == 2 and cell.row > 1:
                        cell.number_format = styles['date_format']
                    
                    if isinstance(cell.value, (int, float)):
                        if cell.column in [5, 7]:
                            cell.number_format = styles['number_format']
                        elif cell.column == 8:
                            cell.number_format = styles['money_format']
                    
                    cell.alignment = Alignment(horizontal='left', vertical='center')

            for col, width in {'A':5, 'B':12, 'C':30, 'D':8, 'E':12, 'F':20, 'G':12, 'H':12, 'I':15}.items():
                ws_coming.column_dimensions[col].width = width

            wb.active = 0
            wb.save(report_filename)

            if os.path.exists(report_filename) and os.path.getsize(report_filename) > 0:
                return True
            return False
            
        except Exception as e:
            print(f"Критическая ошибка при экспорте: {str(e)}")
            import traceback
            traceback.print_exc()
            if os.path.exists(report_filename):
                try:
                    os.remove(report_filename)
                except:
                    pass
            return False